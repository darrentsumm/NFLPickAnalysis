import os
import re
import openpyxl
from dotenv import load_dotenv
from supabase import create_client

# --- CONFIGURATION ---
load_dotenv()
SUPABASE_URL = os.getenv("SUPABASE_URL")
SUPABASE_KEY = os.getenv("SUPABASE_KEY")

if not SUPABASE_URL or not SUPABASE_KEY:
    raise ValueError("Missing Supabase credentials in .env file")

#create a client to access supabase
supabase = create_client(SUPABASE_URL, SUPABASE_KEY)

#Called by the main function to process given workbook and for specified season year
def upload_pool_data(workbook_path, season_year):
    print(f"📂 Opening Workbook: {workbook_path} for Season: {season_year}")
    
    try:
        wb = openpyxl.load_workbook(workbook_path, data_only=True)
    except FileNotFoundError:
        print("❌ Error: File not found.")
        return

    # --- 1. PROCESS USERS (Scan all Week Sheets) ---
    # Old workbooks only showed users that made picks each week, so the function needs to scan all sheets
    print("👤 Processing Users across all sheets...")
    unique_usernames = set()
    
    for sheet in wb.worksheets:
        # process sheets that start with "Week"
        if "week" not in sheet.title.lower():
            continue

        for row in sheet.iter_rows(min_row=8, min_col=3, max_col=3):
            cell = row[0]
            username = cell.value
            if username:
                unique_usernames.add(str(username).strip())

    users_to_upsert = [
        {"username": name, f"season_{season_year}": True} 
        for name in unique_usernames
    ]

    if users_to_upsert:
        try:
            supabase.table("User").upsert(users_to_upsert, on_conflict="username").execute()
            print(f"✅ Upserted {len(users_to_upsert)} unique users found across all sheets.")
        except Exception as e:
            print(f"❌ Error uploading users: {e}")
            return
    else:
        print("⚠️ No users found in any Week sheets.")

    # --- 2. PROCESS WEEKS (Iterate Sheets) ---
    for sheet in wb.worksheets:
        sheet_title = sheet.title.strip()
        week_match = re.search(r"Week\s+(\d+)", sheet_title, re.IGNORECASE)
        if not week_match:
            continue
            
        week_num = int(week_match.group(1))
        print(f"\n📅 Processing Season {season_year} | Week {week_num}...")

        # --- A. IDENTIFY VALID GAME COLUMNS FIRST ---
        valid_game_cols = []
        max_col = sheet.max_column

        for col_idx in range(4, max_col + 1):
            row_8_val = sheet.cell(row=8, column=col_idx).value
            away_team_val = sheet.cell(row=4, column=col_idx).value

            if row_8_val is not None and away_team_val is not None:
                valid_game_cols.append(col_idx)

        if not valid_game_cols:
            print("   ⚠️ No games found in this sheet.")
            continue

        mnf_col_index = valid_game_cols[-1]
        print(f"   🏈 Detected {len(valid_game_cols)} Games. MNF is at Column {mnf_col_index}")

        # --- B. UPSERT GAMES ---
        col_to_game_id = {}

        for col_idx in valid_game_cols:
            is_mnf = (col_idx == mnf_col_index)

            home_score = sheet.cell(row=1, column=col_idx).value
            home_team = sheet.cell(row=2, column=col_idx).value
            home_spread = sheet.cell(row=3, column=col_idx).value
            away_team = sheet.cell(row=4, column=col_idx).value
            away_score = sheet.cell(row=5, column=col_idx).value
            ot_val = sheet.cell(row=7, column=col_idx).value

            try:
                h_score = int(float(home_score)) if home_score is not None else 0
                a_score = int(float(away_score)) if away_score is not None else 0
                spread = float(home_spread) if home_spread is not None else 0.0
                
                adjusted_home = h_score + spread
                tie_spread = (adjusted_home == a_score)
                home_cover = (adjusted_home > a_score)
                is_ot = (ot_val == "OT")

            except (ValueError, TypeError):
                print(f"⚠️ Data error in Col {col_idx}, skipping.")
                continue

            game_payload = {
                "season": season_year,
                "week": week_num,
                "home_team_id": str(home_team).strip(),
                "away_team_id": str(away_team).strip(),
                "home_score": h_score,
                "away_score": a_score,
                "home_spread": spread,
                "home_cover": home_cover,
                "tie_spread": tie_spread,
                "ot": is_ot,
                "mnf": is_mnf
            }

            try:
                res = supabase.table("game").upsert(
                    game_payload, 
                    on_conflict="season,week,home_team_id,away_team_id"
                ).execute()
                
                if res.data:
                    col_to_game_id[col_idx] = res.data[0]['game_id']
            except Exception as e:
                print(f"❌ Error uploading game: {e}")

        # --- C. PROCESS PICKS ---
        picks_to_upsert = []
        
        for row_idx in range(8, sheet.max_row + 1):
            username_val = sheet.cell(row=row_idx, column=3).value
            status_val = sheet.cell(row=row_idx, column=2).value
            
            if not username_val:
                continue

            username = str(username_val).strip()

            status_str = str(status_val).lower() if status_val else ""
            is_late = ("late" in status_str)
            no_picks = ("no picks" in status_str)
            pick_made_row = not no_picks

            for col_idx in valid_game_cols:
                game_id = col_to_game_id.get(col_idx)
                if not game_id: continue

                pick_cell_val = sheet.cell(row=row_idx, column=col_idx).value
                home_team_abbr = str(sheet.cell(row=2, column=col_idx).value).strip()
                
                pick_home = False
                
                if pick_cell_val:
                    user_picked_team = str(pick_cell_val).strip()
                    if user_picked_team == home_team_abbr:
                        pick_home = True

                actual_pick_made = (pick_made_row and (pick_cell_val is not None))

                tot_points = None
                if col_idx == mnf_col_index:
                    target_col = mnf_col_index + 3
                    tot_val = sheet.cell(row=row_idx, column=target_col).value
                    if tot_val:
                        try:
                            tot_points = int(float(tot_val))
                        except:
                            tot_points = None

                picks_to_upsert.append({
                    "username": username,
                    "game_id": game_id,
                    "pick_home": pick_home,
                    "pick_made": actual_pick_made,
                    "pick_overwritten": is_late,
                    "tot_if_picked": tot_points
                })
        
        if picks_to_upsert:
            try:
                supabase.table("pick").upsert(
                    picks_to_upsert, 
                    on_conflict="username,game_id"
                ).execute()
                print(f"   ✅ Processed {len(picks_to_upsert)} picks.")
            except Exception as e:
                print(f"   ❌ Error uploading picks: {e}")

if __name__ == "__main__":
    # Update filename and year as needed
    FILE_NAME = '/Users/darrensummerlee/Documents/Personal Projects/DS_Projects/poolhost/Data/cleaned data/Poolhost 15 edited.xlsx'
    YEAR_INPUT = 2015
    
    upload_pool_data(FILE_NAME, YEAR_INPUT)